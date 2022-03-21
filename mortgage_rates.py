from bs4 import BeautifulSoup as b
from numpy import average
import requests as r
import pandas as p
import openpyxl


excel = openpyxl.Workbook()  # creates an excel file
# print(excel.sheetnames)  # how many sheets are created
sheet = excel.active  # make sure we're working of the main sheet
sheet.title = 'Mortgage_Rates'  # changed the sheet name
# print(excel.sheetnames)
# added a row in the file
sheet.append(['Year', 'Lowest Rate', 'Highest Rate', 'Average Rate'])

# url extracting from
# https://www.hsh.com/monthly-mortgage-rates.html
myUrl = 'https://www.valuepenguin.com/mortgages/historical-mortgage-rates'

# get data from url
download = r.get(myUrl)

# enable to extract the html text
soup = b(download.text, "html.parser")

rates = soup.find(
    'tbody', class_="StyledBody-sc-14y8oc0 dcdMVT").find_all('tr')

# the td had the same class name. The script only pulled the first number. I used split to return a list and then the index of the element
for rate in rates:
    year = (str(rate).replace(
        '<tr><td class="StyledBodyCell-sc-5cu9ee XfhRD" colspan="1" rowspan="1" width="">', '').replace('</td>', '').replace('<td class="StyledBodyCell-sc-5cu9ee frAYUA" colspan="1" rowspan="1" width="">', ' ').replace('</tr>', ' ').strip().split()[0])

    low = (str(rate).replace(
        '<tr><td class="StyledBodyCell-sc-5cu9ee XfhRD" colspan="1" rowspan="1" width="">', '').replace('</td>', '').replace('<td class="StyledBodyCell-sc-5cu9ee frAYUA" colspan="1" rowspan="1" width="">', ' ').replace('</tr>', ' ').strip().split()[1])

    high = (str(rate).replace(
        '<tr><td class="StyledBodyCell-sc-5cu9ee XfhRD" colspan="1" rowspan="1" width="">', '').replace('</td>', '').replace('<td class="StyledBodyCell-sc-5cu9ee frAYUA" colspan="1" rowspan="1" width="">', ' ').replace('</tr>', ' ').strip().split()[2])

    avg = (str(rate).replace(
        '<tr><td class="StyledBodyCell-sc-5cu9ee XfhRD" colspan="1" rowspan="1" width="">', '').replace('</td>', '').replace('<td class="StyledBodyCell-sc-5cu9ee frAYUA" colspan="1" rowspan="1" width="">', ' ').replace('</tr>', ' ').strip().split()[3])

    print(year, low, high, avg)
    sheet.append([year, low, high, avg])

excel.save('Mortgage Rates.xlsx')
