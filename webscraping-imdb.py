from bs4 import BeautifulSoup as b
import requests as r
import pandas as p
import openpyxl

excel = openpyxl.Workbook()  # creates an excel file
print(excel.sheetnames)  # how many sheets are created
sheet = excel.active  # make sure we're working of the main sheet
sheet.title = 'Top Rated Movies'  # changed the sheet name
print(excel.sheetnames)
sheet.append(['Rank', 'Movie Title', 'Year Released',
             'IMDB Ratings'])  # created a row in the file

try:  # this will print error in case the url is not valid and avoid crashing the system
    myUrl = 'https://www.imdb.com/chart/top?pf_rd_m=A2FGELUUNOQJNL&pf_rd_p=470df400-70d9-4f35-bb05-8646a1195842&pf_rd_r=JSRC950SSY27FSK79KS0&pf_rd_s=right-4&pf_rd_t=15506&pf_rd_i=moviemeter&ref_=chtmvm_ql_3'
    download = r.get(myUrl)
    soup = b(download.text, "html.parser")

    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    for mov in movies:
        ranking = mov.find('td', class_="titleColumn").get_text(
            strip=True).split('.')[0]
        title = mov.find('td', class_="titleColumn").a.text
        year = mov.find('td', class_="titleColumn").span.text.strip('()')
        ratings = mov.find('td', class_="ratingColumn imdbRating").strong.text

        #print(ranking, title, year, ratings)
        sheet.append([ranking, title, year, ratings])

except Exception as e:
    print(e)

excel.save('IMDB Top 250 Movies.xlsx')  # save this to a excel file
